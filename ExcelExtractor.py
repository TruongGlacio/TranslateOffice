from openpyxl import load_workbook
from openpyxl.drawing import graphic
import GeneralFuntions

generalFunctions=GeneralFuntions.General()
class ExcelExtrator:

    def __init__(self):
        pass
    
    def ExcelReadChart(self,inputPath):
        excelTextList=list()        
        wb=load_workbook(inputPath)
        
        for ws in wb.chartsheets:
            
            print('sheet column :', ws.max_row) 
    def ExcelReadText(self,inputPath,outPutTextFilePath):
        excelTextList=list()        
        wb=load_workbook(inputPath)
        for ws in wb.worksheets: 
            print('sheet column :', ws.max_row)            
            for row in range(1,ws. max_row+1):
                for col in range(1,ws.max_column+1):
                    celObject=ws.cell(row,col)
                    textList=celObject.value
                    if(textList==None):
                        pass
                    else:
                        excelTextList.append(textList)
                        print('value: ', textList)  
        print('List Count:',excelTextList.count)
        jsonList=generalFunctions.ConvertStringToJson(excelTextList)
        generalFunctions.WriteTextFile(jsonList,outPutTextFilePath)
        return jsonList
    def ExcelWriteText(self,inputPath, inputListJson):
        i=0 
        inputListString=generalFunctions.ConvertJsonToString(inputListJson)
        wb=load_workbook(inputPath)
        for ws in wb.worksheets: 
            print('sheet column :', ws.max_row)            
            for row in range(1,ws. max_row+1):
                for col in range(1,ws.max_column+1):
                    celObject=ws.cell(row,col)
                    textList=celObject.value
                    if(textList==None):
                        pass
                    else:
                        ws.cell(row,col).value=inputListString[i]
                        i=i+1
                        print('Cell value change: ', ws.cell(row,col).value)  
        wb.save(inputPath)        